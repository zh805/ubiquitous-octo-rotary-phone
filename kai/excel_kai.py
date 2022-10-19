# -*- coding: utf-8 -*-
#
# @Time           2022/10/17 4:17 PM
# @File           excel_kai.py
# @Description    excel文件处理
# @Author
#

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
# import pandas as pd
import os


def test_1():
    os.chdir('C:/Users/86138/Desktop')
    path = '照片'
    writer = pd.ExcelWriter("img.xlsx")
    list = []
    for root, dirs, files in os.walk(path):
        for file in files:
            print(file)
            file = file.rstrip(".jpg")
            list.append(file)
    dict = {'filename': list}
    print(dict)
    df = pd.DataFrame(dict)
    df.to_excel(writer, 'sheet1', startcol=0, index=False)
    writer.save()


def test_2():
    # 文件路径; 换成自己本地的文件路径
    file = "/Users/zhanghui/Downloads/坐标.xlsx"
    # 使用pandas读文件
    df = pd.read_excel(file, engine='openpyxl')

    # new_name存提取出来的坐标
    new_name = []
    # 一行一行读excel文件
    for idx, row in df.iterrows():
        # 读filename这一列
        name = row["filename"]
        print(f'这一行的文件名是: {name}')

        # 使用下划线 _ 分割字符串，name分割后的结果是一个列表。 比如： ['ADE1712A115', '-79396', '-191018', 'FIXREVIEW.JPG']
        a = name.split('_')
        print(f'分割后的结果是:{a}')

        # lstrip: 把字符串左边的 - 去掉。
        # 把列表中的坐标取出来然后去掉 -
        x = a[1].lstrip('-')
        y = a[2].lstrip('-')
        print(f'提取出的坐标x是:{x},坐标y是{y}')

        # 把 x和y使用逗号拼成一个字符串
        coordinate = x + ',' + y
        print(f'拼接后的坐标是:{coordinate}')

        # 把拼好的字符串放入new_name列表
        new_name.append(coordinate)

    print(f'所有提取出的坐标是: {new_name}')

    # 把提取出来的坐标写到一个新文件
    data = {"坐标": new_name}
    # 新文件的路径；换成自己电脑的路径。
    new_file = "/Users/zhanghui/Downloads/新坐标.xlsx"
    print(f'开始把数据写到新文件: {new_file}')

    # 使用pandas把数据写到新文件
    writer = pd.ExcelWriter(new_file)
    df = pd.DataFrame(data)
    df.to_excel(writer, 'sheet1', startcol=0, index=False)
    writer.save()
    print("数据已写入新文件")


def test_3():
    # 文件路径; 换成自己本地的文件路径
    filename = '坐标.xlsx'
    filepath = os.path.join(os.path.dirname(__file__), filename)
    # 使用pandas读文件
    df = pd.read_excel(filepath, engine='openpyxl')

    # 分别保存提取到的文件名、坐标x，坐标y
    filenames, x_list, y_list = [], [], []
    # 一行一行读excel文件
    for idx, row in df.iterrows():
        # 读filename这一列
        name = row["filename"]
        print(f'这一行的文件名是: {name}')
        filenames.append(name)

        # 使用下划线 _ 分割字符串，name分割后的结果是一个列表。 比如： ['ADE1712A115', '-79396', '-191018', 'FIXREVIEW.JPG']
        a = name.split('_')
        print(f'分割后的结果是:{a}')

        # lstrip: 把字符串左边的 - 去掉。
        # 把列表中的坐标取出来然后去掉 -
        x = a[1].lstrip('-')
        y = a[2].lstrip('-')
        print(f'提取出的坐标x是:{x},坐标y是{y}')

        # 把x,y坐标放入列表中保存
        x_list.append(int(x))
        y_list.append(int(y))

    # 把提取出来的数据写到一个新文件，一共三列。
    data = {"filename": filenames, "坐标x": x_list, "坐标y": y_list}
    # 新文件的路径；换成自己电脑的路径。
    new_file = os.path.join(os.path.dirname(__file__), "新坐标.xlsx")
    print(f'开始把数据写到新文件: {new_file}')

    # 使用pandas把数据写到新文件
    writer = pd.ExcelWriter(new_file)
    df = pd.DataFrame(data)
    df.to_excel(writer, 'sheet1', startcol=0, index=False)
    writer.save()
    print("数据已写入新文件")


def test_4(filename):
    # 文件路径; 换成自己本地的文件路径
    filepath = os.path.join(os.path.dirname(__file__), filename)

    # 使用openpyxl读文件
    wb = load_workbook(filepath)
    ws = wb.active

    # 分别保存提取到的文件名、坐标x，坐标y
    new_rows = [["filename", "坐标x", "坐标y"]]
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        # 读filename这一列
        name = row[0]
        print(f'这一行的文件名是: {name}')

        # 使用下划线 _ 分割字符串，name分割后的结果是一个列表。 比如： ['ADE1712A115', '-79396', '-191018', 'FIXREVIEW.JPG']
        a = name.split('_')
        print(f'分割后的结果是:{a}')

        # lstrip: 把字符串左边的 - 去掉。
        # 把列表中的坐标取出来然后去掉 -
        x = a[1].lstrip('-')
        y = a[2].lstrip('-')
        print(f'提取出的坐标x是:{x},坐标y是{y}')

        new_rows.append([name, int(x), int(y)])

    # 新建excel文件写入
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'sheet1'
    for row in new_rows:
        ws1.append(row)

    # 设置散点图
    chart = ScatterChart()
    chart.style = 10
    chart.title = "AOI点位图"
    chart.x_axis.title = 'X'
    chart.y_axis.title = 'Y'

    # 设置图表大小
    chart.height = 30
    chart.width = 50

    # x，y轴坐标范围
    chart.x_axis.scaling.min = -365000
    chart.x_axis.scaling.max = 365000
    chart.y_axis.scaling.min = -460000
    chart.y_axis.scaling.max = 460000

    # 不加图例
    chart.legend = None

    xdata = Reference(ws1, min_row=2, max_row=len(new_rows), min_col=2, max_col=2)
    ydata = Reference(ws1, min_row=2, max_row=len(new_rows), min_col=3, max_col=3)
    series1 = Series(ydata, xdata, title_from_data=False)
    # 设置系列1数据点的样式，圆圈
    series1.marker.symbol = "circle"
    # 设置系列1数据点的颜色，以下两行代码将其改为蓝色
    series1.marker.graphicalProperties.solidFill = "1E90FF"  # 点的内部填充颜色
    series1.marker.graphicalProperties.line.solidFill = "1E90FF"  # 点的外边框颜色
    # 关闭系列1数据点之间的连接线
    series1.graphicalProperties.line.noFill = True
    chart.series.append(series1)

    # 绘制图表的位置
    ws1.add_chart(chart, 'E' + str(len(new_rows) + 2))

    # 在同一文件夹下生成新文件
    new_file = os.path.join(os.path.dirname(__file__), "新坐标.xlsx")
    print(f'开始把数据写到新文件: {new_file}')
    wb.save(new_file)
    print("数据处理与图表绘制完毕")


def run():
    cur_dir = os.path.dirname(__file__)

    new_rows = [["filename", "坐标x", "坐标y"]]
    for root, dirs, files in os.walk(cur_dir):
        for file in files:
            if 'JPG' not in file:
                continue
            print(file)
            name = file.rstrip(".JPG")

            # 使用下划线 _ 分割字符串，name分割后的结果是一个列表。 比如： ['ADE1712A115', '-79396', '-191018', 'FIXREVIEW.JPG']
            a = name.split('_')
            print(f'分割后的结果是:{a}')

            # lstrip: 把字符串左边的 - 去掉。
            # 把列表中的坐标取出来然后去掉 -
            x = a[1].lstrip('-')
            y = a[2].lstrip('-')
            print(f'提取出的坐标x是:{x},坐标y是{y}')

            new_rows.append([name, int(x), int(y)])

    # 新建excel文件写入
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'sheet1'
    for row in new_rows:
        ws1.append(row)

    # 设置散点图
    chart = ScatterChart()
    chart.style = 10
    chart.title = "AOI点位图"
    chart.x_axis.title = 'X'
    chart.y_axis.title = 'Y'

    # 设置图表大小
    chart.height = 30
    chart.width = 50

    # x，y轴坐标范围
    chart.x_axis.scaling.min = -365000
    chart.x_axis.scaling.max = 365000
    chart.y_axis.scaling.min = -460000
    chart.y_axis.scaling.max = 460000

    # 不加图例
    chart.legend = None

    xdata = Reference(ws1, min_row=2, max_row=len(new_rows), min_col=2, max_col=2)
    ydata = Reference(ws1, min_row=2, max_row=len(new_rows), min_col=3, max_col=3)
    series1 = Series(ydata, xdata, title_from_data=False)
    # 设置系列1数据点的样式，圆圈
    series1.marker.symbol = "circle"
    # 设置系列1数据点的颜色，以下两行代码将其改为蓝色
    series1.marker.graphicalProperties.solidFill = "1E90FF"  # 点的内部填充颜色
    series1.marker.graphicalProperties.line.solidFill = "1E90FF"  # 点的外边框颜色
    # 关闭系列1数据点之间的连接线
    series1.graphicalProperties.line.noFill = True
    chart.series.append(series1)

    # 绘制图表的位置
    ws1.add_chart(chart, 'E' + str(len(new_rows) + 2))

    # 在同一文件夹下生成新文件
    new_file = os.path.join(cur_dir, "AOI.xlsx")
    print(f'开始把数据写到新文件: {new_file}')
    wb.save(new_file)
    print("数据处理与图表绘制完毕")


if __name__ == '__main__':
    # test_1()
    # test_2()
    # test_3()

    # 文件名设置成真实文件的名字
    # filename = "坐标.xlsx"
    # test_4(filename=filename)
    run()
